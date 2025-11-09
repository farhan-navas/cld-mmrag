import json
import pandas as pd

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

JSON_PATH = "data/validation/answer-list.json"
EXCEL_PATH = "data/answer-list.xlsx"

def json_to_pretty_excel(json_path: str, excel_path: str) -> None:
    # 1. Load JSON
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    df = pd.DataFrame(data)
    df = df.where(pd.notnull(df), "")  # replace None with ""

    # Desired column order (optional)
    desired_cols = [
        "sn",
        "query",
        "expected_answer",
        "layer",
        "source_file",
        "actual_answer",
        "citations",
        "follow_up",
        "response_time_ms",
        "success",
        "error",
    ]
    cols_present = [c for c in desired_cols if c in df.columns]
    df = df[cols_present]

    # 2. Write to Excel with openpyxl engine
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        sheet_name = "QA"
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        # Columns that usually contain long text
        wrap_cols = ["query", "expected_answer", "actual_answer"]

        # Map column names -> Excel column letters
        col_name_to_letter = {}
        for idx, col_name in enumerate(df.columns, start=1):
            col_name_to_letter[col_name] = get_column_letter(idx)

        # 3. Enable wrap_text + top alignment for selected columns
        for col_name in wrap_cols:
            if col_name not in col_name_to_letter:
                continue  # skip if column doesn't exist
            col_letter = col_name_to_letter[col_name]

            # Set a nice column width
            ws.column_dimensions[col_letter].width = 40

            # Apply alignment to all data cells in that column
            for row in range(2, len(df) + 2):  # row 1 is header
                cell = ws[f"{col_letter}{row}"]
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # 4. (Optional) make all data rows a bit taller
        for row in range(2, len(df) + 2):
            ws.row_dimensions[row].height = 40

    print(f"Saved formatted Excel file to {excel_path}")

if __name__ == "__main__":
    json_to_pretty_excel(JSON_PATH, EXCEL_PATH)
